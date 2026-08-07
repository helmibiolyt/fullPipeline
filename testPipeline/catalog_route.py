#!/usr/bin/env python3
"""Which of the 134 catalog questions these two stores can actually answer.

    python testPipeline/catalog_route.py            summary
    python testPipeline/catalog_route.py --list     every question with its verdict

Routing is a question about where an answer LIVES. Before choosing between the
graph and the vector store, the catalog has to be split by whether the answer
is in either of them - and a large part of it is not. Data Management asks
about open queries, edit-check failures and database lock; Clinical Leadership
asks about budget, milestones and CRO delivery. Those live in the sponsor's
EDC, CTMS and safety database. No routing rule reaches them, and an agent that
tries will either say nothing or invent something.

So each question gets one of:

    GRAPH      counts, filters, aggregations over trials/drugs/approvals/AEs
    DOCS       wording, findings, narrative - what a document SAYS
    BOTH       needs a structured set AND what the documents say about it
    ABSENT     the answer is in a system this pipeline does not hold

The split is by persona with per-question exceptions, because persona predicts
it well: Competitive Intelligence is almost entirely public registry and label
data, Data Management is almost entirely internal.
"""
from __future__ import annotations

import argparse
import collections
import json
import pathlib
import re
import sys

HERE = pathlib.Path(__file__).resolve().parent
XLSX = HERE / "biolytai_question_catalog_expanded.xlsx"

# Systems this pipeline does not hold: the sponsor's own operational data.
_ABSENT = re.compile(
    r"\b(open quer|queries|discrepanc|edit.check|database lock|CRF|form\b|forms\b|"
    r"medical coding backlog|data entry|SAE reconciliation|missingness|"
    r"pending medical review|15-day|reporting deadline|case status|workflow|"
    r"screening activity|screen.fail|protocol deviation|enrollment target|"
    r"enrolled subjects|milestone|budget|resource allocation|CRO|vendor|"
    r"monitoring visit|site activation|startup|randomi[sz]ation|"
    r"submission status|regulatory commitment|health authority questions received|"
    r"KOL interactions|resourcing|timeline slippage|risk count)\b", re.I)

# Wording, findings, narrative - what a document says rather than how many.
_DOCS = re.compile(
    r"\b(evidence|publication|publish|abstract|congress|narrative|draft|"
    r"summar|dossier|brief|label(ing)? claim|claim\b|guidance|precedent|"
    r"real.world|RWE|mechanism.of.action|scientific (question|response)|"
    r"value narrative|positioning|overview)\b", re.I)

# Counting, filtering, ranking, grouping over structured records.
_GRAPH = re.compile(
    r"\b(list|show|how many|count|rank|compare|map the|pipeline|phase|status|"
    r"sponsor|approved|approval|trial|registry|indication|by (phase|mechanism|"
    r"sponsor|geography|System Organ Class|SOC)|disproportionality|signal|"
    r"co-reported|trend|volume)\b", re.I)


def verdict(q: str, corpus: str, persona: str) -> str:
    blob = f"{q} {corpus}"
    if _ABSENT.search(blob):
        return "ABSENT"
    # Three personas are operational by definition, and the keyword pass is not
    # enough on its own: it read "Show open protocol deviations by site" and
    # "enrollment trajectory vs plan" as answerable because they contain "show"
    # and "site". Every question in these three sheets is about the sponsor's
    # own study conduct - its sites, its subjects, its forms, its budget - and
    # none of that is in a public registry or a document corpus.
    #
    # Clinical Operations was the one I got wrong first time round: it scored
    # 13 of 20 answerable on keywords alone, and reading them showed the true
    # figure is nearer zero.
    if persona in ("Data Management", "Clinical Leadership",
                   "Clinical Operations"):
        return "ABSENT"
    d, g = bool(_DOCS.search(blob)), bool(_GRAPH.search(blob))
    return "BOTH" if d and g else "DOCS" if d else "GRAPH" if g else "BOTH"


def load() -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    out = []
    for name in wb.sheetnames:
        if name == "Overview":
            continue
        rows = list(wb[name].iter_rows(values_only=True))
        h = next((i for i, r in enumerate(rows) if r and r[0] == "ID"), None)
        if h is None:
            continue
        cols = [str(c).strip() if c else "" for c in rows[h]]
        for r in rows[h + 1:]:
            if not r or not r[0]:
                continue
            d = {cols[i]: (str(r[i]).strip() if r[i] is not None else "")
                 for i in range(min(len(cols), len(r)))}
            d["persona"] = name
            d["question"] = d.get("Persona Question (card label)", "")
            d["verdict"] = verdict(d["question"],
                                   d.get("Corpus Data Leveraged", ""), name)
            out.append(d)
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--list", action="store_true")
    ap.add_argument("--json", type=pathlib.Path)
    a = ap.parse_args()

    qs = load()
    by_p = collections.defaultdict(collections.Counter)
    tot = collections.Counter()
    for q in qs:
        by_p[q["persona"]][q["verdict"]] += 1
        tot[q["verdict"]] += 1

    order = ["GRAPH", "DOCS", "BOTH", "ABSENT"]
    print(f"{len(qs)} questions\n")
    print(f"{'persona':<26}" + "".join(f"{k:>8}" for k in order) + f"{'answerable':>12}")
    print("-" * 82)
    for p in sorted(by_p, key=lambda x: -sum(by_p[x][k] for k in order[:3])):
        c = by_p[p]
        ans = sum(c[k] for k in order[:3])
        print(f"{p:<26}" + "".join(f"{c[k]:>8}" for k in order)
              + f"{ans:>7} /{sum(c.values()):>3}")
    print("-" * 82)
    ans = sum(tot[k] for k in order[:3])
    print(f"{'TOTAL':<26}" + "".join(f"{tot[k]:>8}" for k in order)
          + f"{ans:>7} /{len(qs):>3}")
    print(f"\nanswerable from these two stores: {ans}/{len(qs)} ({ans/len(qs):.0%})")
    print(f"needs the sponsor's own systems:  {tot['ABSENT']}/{len(qs)} "
          f"({tot['ABSENT']/len(qs):.0%})")

    if a.list:
        print()
        for q in sorted(qs, key=lambda x: (x["verdict"], x["persona"])):
            print(f"  {q['verdict']:<7} [{q['persona'][:20]:<20}] {q['question'][:70]}")
    if a.json:
        a.json.write_text(json.dumps(qs, ensure_ascii=False, indent=1),
                          encoding="utf-8")
        print(f"\nwrote {a.json}")


if __name__ == "__main__":
    sys.exit(main())
